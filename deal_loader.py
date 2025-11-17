from __future__ import annotations

import argparse
import json
import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Tuple

try:
    from openpyxl import Workbook, load_workbook
except ImportError as exc:  # pragma: no cover
    raise SystemExit("Please install openpyxl to run the loader (pip install openpyxl).") from exc

try:
    import win32com.client as win32  # type: ignore
except ImportError:  # pragma: no cover
    win32 = None


@dataclass
class DealResult:
    spv: str
    file_path: Path
    business_date: str
    static_values: Dict[str, Any]
    cell_values: Dict[str, Any]
    calculated_fields: Dict[str, Any]
    combined_values: Dict[str, Any]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Load Excel-based deal configs and draft an Outlook summary."
    )
    parser.add_argument(
        "--config-dir",
        default="configs",
        help="Directory containing deal config JSON files (default: configs)",
    )
    parser.add_argument(
        "--sample-data",
        default="sample_data",
        help="Directory where demo Excel files will be created (default: sample_data)",
    )
    parser.add_argument(
        "--force-demo",
        action="store_true",
        help="Rebuild the demo Excel file even if it already exists.",
    )
    parser.add_argument(
        "--subject",
        default="Deal Loader Report (Draft)",
        help="Subject line for the Outlook draft email.",
    )
    parser.add_argument(
        "--report-profile",
        default="report_profile.json",
        help="JSON file that defines the shared summary fields and default detail sections for the report.",
    )
    return parser.parse_args()


def load_configs(config_dir: Path) -> List[Dict[str, Any]]:
    configs: List[Dict[str, Any]] = []
    for config_file in sorted(config_dir.glob("*.json")):
        with config_file.open("r", encoding="utf-8-sig") as handle:
            data = json.load(handle)
        data["_file"] = config_file
        configs.append(data)
    if not configs:
        raise FileNotFoundError(f"No configs found in {config_dir}")
    return configs


DEFAULT_REPORT_PROFILE = {
    "summary_fields": [
        {"label": "Client Name", "source": "client_name"},
        {"label": "Borrower Name", "source": "borrower_name"},
        {"label": "Risk Rating", "source": "risk_rating"},
        {"label": "Closing Date", "source": "closing_date"},
        {"label": "Revolving Period End Date", "source": "revolving_period_end_date"},
        {"label": "Amortization Period End Date", "source": "amortization_period_end_date"},
        {"label": "Facility Maturity Date", "source": "facility_maturity_date"},
        {"label": "As of Date", "source": "@business_date"},
        {"label": "Global Advances Outstanding", "source": "global_advances_outstanding", "format": "currency", "aggregate": "sum"},
        {"label": "Global Commitment", "source": "global_commitment", "format": "currency", "aggregate": "sum"},
        {"label": "BMO Advances Outstanding", "source": "bmo_advances_outstanding", "format": "currency", "aggregate": "sum"},
        {"label": "BMO Commitment", "source": "bmo_commitment", "format": "currency", "aggregate": "sum"},
        {"label": "BMO Utilization", "source": "bmo_utilization", "format": "percentage", "aggregate": "ratio", "numerator": "bmo_advances_outstanding", "denominator": "bmo_commitment"},
        {"label": "Collateral Par Balance", "source": "collateral_par_balance", "format": "currency", "aggregate": "sum"},
        {"label": "Max Facility Advance Rate", "source": "max_facility_advance_rate", "format": "percentage"},
        {"label": "Effective Advance Rate", "source": "effective_advance_rate", "format": "percentage", "aggregate": "ratio", "numerator": "global_advances_outstanding", "denominator": "collateral_par_balance"}
    ],
    "detail_defaults": [
        {
            "title": "Deal Dates",
            "fields": [
                {"label": "Closing Date", "source": "closing_date"},
                {"label": "Revolving Period End", "source": "revolving_period_end_date"},
                {"label": "Amortization Period End", "source": "amortization_period_end_date"},
                {"label": "Facility Maturity", "source": "facility_maturity_date"},
            ],
        },
        {
            "title": "Advances & Commitments",
            "fields": [
                {"label": "Global Advances Outstanding", "source": "global_advances_outstanding", "format": "currency"},
                {"label": "Global Commitment", "source": "global_commitment", "format": "currency"},
                {"label": "BMO Advances Outstanding", "source": "bmo_advances_outstanding", "format": "currency"},
                {"label": "BMO Commitment", "source": "bmo_commitment", "format": "currency"},
                {"label": "Collateral Par Balance", "source": "collateral_par_balance", "format": "currency"},
            ],
        },
        {
            "title": "Rates",
            "fields": [
                {"label": "BMO Utilization", "source": "bmo_utilization", "format": "percentage"},
                {"label": "Max Facility Advance Rate", "source": "max_facility_advance_rate", "format": "percentage"},
                {"label": "Effective Advance Rate", "source": "effective_advance_rate", "format": "percentage"},
            ],
        },
    ]
}


def load_report_profile(path: Path) -> Dict[str, Any]:
    if path.exists():
        with path.open("r", encoding="utf-8-sig") as handle:
            return json.load(handle)
    return DEFAULT_REPORT_PROFILE


def sanitize_name(value: str) -> str:
    return re.sub(r"[^A-Za-z0-9]+", "_", value).strip("_").lower() or "deal"


def ensure_demo_workbook(config: Dict[str, Any], sample_dir: Path, force: bool = False) -> Path:
    sample_dir.mkdir(parents=True, exist_ok=True)
    filename = f"{sanitize_name(config['spv'])}_demo.xlsx"
    destination = sample_dir / filename
    if destination.exists() and not force:
        return destination

    wb = Workbook()
    ws = wb.active
    ws.title = "Capital Structure"

    ws["A2"] = "Business Date"
    ws["B2"] = datetime(2025, 10, 11)
    ws["A4"] = "Global Commitment"
    ws["B4"] = 12_500_000
    ws["B5"] = 10_000_000
    ws["B10"] = 8_000_000

    ws["A12"] = "Notes"
    ws["B12"] = "Demo workbook generated by loader."

    wb.save(destination)
    return destination


def find_matching_files(config: Dict[str, Any], search_dir: Path) -> List[Path]:
    pattern = config.get("file_pattern", "")
    regex = re.compile(pattern)
    return [path for path in search_dir.glob("*.xlsx") if regex.search(path.name)]


def parse_sheet_cell(value: str) -> Tuple[str, str]:
    parts = [part.strip() for part in value.split(",")]
    if len(parts) != 2:
        raise ValueError(f"Expected 'Sheet, Cell' format, got: {value}")
    return parts[0], parts[1]


def extract_cell(workbook, sheet: str, cell: str):
    if sheet not in workbook.sheetnames:
        raise ValueError(f"Sheet '{sheet}' not found in {workbook.sheetnames}")
    ws = workbook[sheet]
    return ws[cell].value


def safe_number(value: Any) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).replace(",", ""))
    except ValueError:
        return 0.0


def evaluate_calculations(calculated: Dict[str, Any], context: Dict[str, Any]) -> Dict[str, Any]:
    results: Dict[str, Any] = {}
    safe_locals = {key: safe_number(val) for key, val in context.items()}
    for key, spec in calculated.items():
        formula = spec.get("formula", "")
        try:
            value = eval(  # noqa: S307 - controlled input from configs
                formula,
                {"__builtins__": {}},
                safe_locals,
            )
        except Exception as exc:
            value = f"Error: {exc}"
        results[key] = value
    return results


def process_file(config: Dict[str, Any], workbook_path: Path) -> DealResult:
    wb = load_workbook(workbook_path, data_only=True)
    static_values = config["fields"].get("static_values", {}).copy()

    cell_values: Dict[str, Any] = {}
    for name, location in config["fields"].get("cell_references", {}).items():
        sheet = location.get("sheet", "")
        cell = location.get("cell", "")
        if sheet and cell:
            cell_values[name] = extract_cell(wb, sheet, cell)

    variable_values: Dict[str, Any] = {}
    for name, location in config["fields"].get("variables", {}).items():
        sheet = location.get("sheet", "")
        cell = location.get("cell", "")
        if sheet and cell:
            variable_values[name] = extract_cell(wb, sheet, cell)

    calculation_context = {**static_values, **cell_values, **variable_values}
    calculated_values = evaluate_calculations(
        config["fields"].get("calculated_fields", {}),
        calculation_context,
    )

    business_date = extract_business_date(config, workbook_path, wb)

    combined_values = {
        **static_values,
        **cell_values,
        **variable_values,
        **calculated_values,
    }

    return DealResult(
        spv=config["spv"],
        file_path=workbook_path,
        business_date=business_date,
        static_values=static_values,
        cell_values=cell_values,
        calculated_fields=calculated_values,
        combined_values=combined_values,
    )


def extract_business_date(config: Dict[str, Any], workbook_path: Path, workbook) -> str:
    ds = config.get("data_source", {})
    source_type = ds.get("type", "filename")
    pattern = ds.get("regex", "")
    if source_type == "cell_reference" and pattern:
        sheet, cell = parse_sheet_cell(pattern)
        value = extract_cell(workbook, sheet, cell)
        if isinstance(value, datetime):
            return value.strftime("%Y-%m-%d")
        return str(value)
    if pattern:
        regex = re.compile(pattern)
        match = regex.search(workbook_path.name)
        if match:
            return match.group(0)
    return datetime.now().strftime("%Y-%m-%d")


SUMMARY_METRICS = [
    ("Total NAV", "nav_total"),
    ("Total Commitment", "total_commitment"),
    ("Capital Called", "capital_called"),
    ("Outstanding Debt", "outstanding_debt"),
    ("Cash on Hand", "cash_on_hand"),
]


def format_value(value: Any) -> str:
    if isinstance(value, (int, float)):
        return f"{value:,.2f}"
    return str(value)


def resolve_source(result: DealResult, source: str) -> Any:
    if source == "@spv":
        return result.spv
    if source == "@file":
        return result.file_path.name
    if source == "@business_date":
        return result.business_date
    return result.combined_values.get(source)


def format_value(value: Any, fmt: str | None = None) -> str:
    if value in (None, "", "—"):
        return "—"
    if fmt == "currency":
        return f"{safe_number(value):,.2f}"
    if fmt == "number":
        return f"{safe_number(value):,.2f}"
    if fmt == "percentage":
        return f"{safe_number(value) * 100:,.2f}%"
    return str(value)


def aggregate_values(values: List[float], method: str) -> float | None:
    if not values:
        return None
    if method == "average":
        return sum(values) / len(values)
    if method == "max":
        return max(values)
    if method == "min":
        return min(values)
    return sum(values)


def render_email(results: List[DealResult], profile: Dict[str, Any]) -> str:
    if not results:
        return "<p>No deals were processed.</p>"

    summary_columns = profile.get("summary_fields", [])
    summary_rows = []
    for result in results:
        cells = []
        for column in summary_columns:
            value = resolve_source(result, column.get("source", ""))
            display = format_value(value, column.get("format"))
            cells.append(f"<td>{display}</td>")
        summary_rows.append("<tr>" + "".join(cells) + "</tr>")

    totals_map: Dict[str, Any] = {}
    for column in summary_columns:
        aggregate_method = column.get("aggregate")
        if not aggregate_method:
            continue
        source = column.get("source", "")
        if aggregate_method == "sum":
            totals_map[source] = sum(
                safe_number(resolve_source(result, source)) for result in results
            )
        elif aggregate_method == "ratio":
            numerator_field = column.get("numerator") or source
            denominator_field = column.get("denominator")
            numerator = sum(safe_number(resolve_source(result, numerator_field)) for result in results)
            denominator = sum(
                safe_number(resolve_source(result, denominator_field)) for result in results
            ) if denominator_field else 0
            totals_map[source] = numerator / denominator if denominator else None
        elif aggregate_method == "average":
            values = [
                safe_number(resolve_source(result, source))
                for result in results
                if resolve_source(result, source) not in (None, "", "—")
            ]
            totals_map[source] = (sum(values) / len(values)) if values else None
        elif aggregate_method == "max":
            totals_map[source] = max(
                safe_number(resolve_source(result, source)) for result in results
            )
        elif aggregate_method == "min":
            totals_map[source] = min(
                safe_number(resolve_source(result, source)) for result in results
            )

    total_row_cells = []
    for idx, column in enumerate(summary_columns):
        if idx == 0:
            total_row_cells.append("<td><strong>TOTAL</strong></td>")
            continue
        column_source = column.get("source", "")
        aggregate_value = totals_map.get(column_source)
        display = format_value(aggregate_value, column.get("format")) if aggregate_value is not None else ""
        total_row_cells.append(f"<td class='total-cell'>{display}</td>")

    summary_table = [
        "<table class='summary-table'>",
        "<tr>" + "".join(f"<th>{column.get('label')}</th>" for column in summary_columns) + "</tr>",
        *summary_rows,
        "<tr class='total-row'>" + "".join(total_row_cells) + "</tr>",
        "</table>",
    ]

    # Detail sections per deal
    detail_sections = []
    detail_rows_def = profile.get("detail_rows", [])
    for result in results:
        rows_html = []
        if detail_rows_def:
            for row in detail_rows_def:
                left_label = row.get("left_label", "")
                left_type = row.get("left_type") or "field"
                left_value = (
                    row.get("left_text", "")
                    if left_type == "text"
                    else format_value(resolve_source(result, row.get("left_source", "")))
                )
                right_label = row.get("right_label", "")
                right_type = row.get("right_type") or "field"
                right_value = (
                    row.get("right_text", "")
                    if right_type == "text"
                    else format_value(resolve_source(result, row.get("right_source", "")))
                )
                rows_html.append(
                    "<tr>"
                    f"<th>{left_label}</th><td>{left_value}</td>"
                    f"<th>{right_label}</th><td>{right_value}</td>"
                    "</tr>"
                )
        else:
            rows_html.append("<tr><td colspan='4'>No detail rows configured.</td></tr>")

        detail_sections.append(
            f"""
            <div class='deal-card'>
                <div class='deal-header'>
                    <div><h3>{result.spv}</h3></div>
                    <div class='deal-date'><strong>As of:</strong> {result.business_date}</div>
                </div>
                <table class='detail-grid'>
                    {''.join(rows_html)}
                </table>
            </div>
            """
        )

    html = f"""
    <style>
        body {{ font-family: 'Segoe UI', Arial, sans-serif; background:#f5f7fb; color:#0f1a33; }}
        .wrapper {{ max-width: 960px; margin: 0 auto; padding: 20px; }}
        .summary-table {{ width:100%; border-collapse: collapse; margin-bottom: 32px; background:#fff; border-radius: 16px; overflow:hidden; box-shadow:0 8px 24px rgba(15,42,99,0.08); }}
        .summary-table th {{ background:#1a3bb5; color:#fff; padding:10px; text-align:left; }}
        .summary-table td {{ padding:10px; border-bottom:1px solid #eef2fb; }}
        .summary-table .total-row td {{ background:#f0f3ff; font-weight:600; }}
        .deal-card {{ background:#fff; border-radius: 16px; padding: 20px; margin-bottom: 18px; box-shadow:0 8px 24px rgba(15,42,99,0.08); }}
        .deal-header {{ display:flex; justify-content:space-between; align-items:flex-start; margin-bottom:12px; }}
        .deal-header h3 {{ margin:0; }}
        .deal-header p {{ margin:4px 0 0; color:#5a6b8c; }}
        .detail-grid {{ width:100%; border-collapse: collapse; margin-top:10px; }}
        .detail-grid th {{ background:#f5f7ff; text-transform:uppercase; font-size:0.75rem; letter-spacing:0.05em; padding:8px; color:#4d6090; }}
        .detail-grid td {{ padding:8px; border-bottom:1px solid #eef2fb; font-weight:600; color:#16255a; }}
    </style>
    <div class="wrapper">
        <h2>Deal Summary</h2>
        {''.join(summary_table)}
        <h2>Deal Details</h2>
        {''.join(detail_sections)}
    </div>
    """
    return html


def open_outlook_draft(html_body: str, subject: str) -> None:
    if win32 is None:
        print("pywin32 is not installed; skipping Outlook draft creation.")
        return
    try:
        outlook = win32.Dispatch("Outlook.Application")
        mail = outlook.CreateItem(0)
        mail.Subject = subject
        mail.HTMLBody = html_body
        mail.Display()
        print("Outlook draft opened.")
    except Exception as exc:  # pragma: no cover
        print(f"Unable to open Outlook draft: {exc}")


def main() -> None:
    args = parse_args()
    config_dir = Path(args.config_dir)
    sample_dir = Path(args.sample_data)

    configs = load_configs(config_dir)
    results: List[DealResult] = []

    for config in configs:
        demo_workbook = ensure_demo_workbook(config, sample_dir, force=args.force_demo)
        matches = find_matching_files(config, sample_dir)
        if not matches:
            matches = [demo_workbook]
        for workbook_path in matches:
            try:
                result = process_file(config, workbook_path)
            except Exception as exc:
                print(f"Failed to process {workbook_path}: {exc}")
                continue
            results.append(result)

    profile = load_report_profile(Path(args.report_profile))
    html_body = render_email(results, profile)
    open_outlook_draft(html_body, args.subject)
    Path("deal_loader_report.html").write_text(html_body, encoding="utf-8")
    print("Report HTML saved to deal_loader_report.html")


if __name__ == "__main__":
    main()
